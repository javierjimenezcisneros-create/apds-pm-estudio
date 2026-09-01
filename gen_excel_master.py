import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation
import datetime

wb = openpyxl.Workbook()
wb.remove(wb.active)

# ── Paleta APDS (idéntica a los PDFs) ────────────────────────────────────────
C = {
    'granate':  'FF5C1831', 'naranja':  'FFDB7733', 'paper':    'FFF4F1EC',
    'blanco':   'FFFFFFFF', 'gris':     'FF4A4A4A', 'gris_l':   'FF9C9088',
    'verde_d':  'FF1E5C1E', 'verde_l':  'FFD5F5E3', 'rojo_d':   'FFC0392B',
    'rojo_l':   'FFFADBD8', 'azul_d':   'FF1A3A5C', 'azul_l':   'FFD6EAF8',
    'naranja_l':'FFFAE5D0', 'gris_h':   'FFEDE8E0',
    # Colores de fase (= PDFs)
    'PROY B':    'FFBDD7EE', 'PROY E':   'FF70B0D8', 'PROY DECO':'FFC9B8E8',
    'ANT.PROY':  'FFD4C4F0', 'OBRA':     'FFF5C07A', 'FIN OBRA': 'FF2E7D32',
    'PEDIDOS':   'FFFFE384', 'MONTAJE':  'FFE87878', 'REUNIÓN':  'FFA8D8A8',
    'ENTREGA':   'FF5FAD5F', 'APERTURA': 'FF5FAD5F', 'REMATES':  'FFF5C07A',
    'VIAJE':     'FF8B2252', 'REVISIÓN': 'FFBDD7EE', 'RENDERS':  'FFC8C8C8',
    'PDTE LIC':  'FFD0CEC8', 'PENDIENTE':'FFEBEBEB', 'FUERA':    'FF888888',
    'CIERRE':    'FFA8D8A8', 'VISITA':   'FFA8D8A8', 'ALEJANDRA':'FF8B2252',
}
FASE_TXT = {  # texto blanco cuando fondo oscuro
    'PROY E','FIN OBRA','FUERA','VIAJE','ALEJANDRA','ENTREGA','APERTURA',
}

def fl(h): return PatternFill('solid', fgColor=_hex6(str(h)))
def _hex6(h):
    h = str(h).strip().upper().lstrip('#')
    if len(h) == 8 and h[:2] == 'FF': h = h[2:]
    if len(h) != 6: h = '4A4A4A'
    return h

def fn(bold=False, color='4A4A4A', size=9, italic=False):
    return Font(bold=bold, color=_hex6(color), size=size, italic=italic, name='Arial')
def al(h='left', v='center', wrap=True):
    return Alignment(horizontal=h, vertical=v, wrap_text=wrap)
def sd(c='D4C5B0'): return Side(style='thin', color=c)
def bd(): s=sd(); return Border(left=s,right=s,top=s,bottom=s)
def bd_none(): n=Side(style=None); return Border(left=n,right=n,top=n,bottom=n)

SEMANAS = [
    ('AGO','1 SEP'),('SEP','8 SEP'),('SEP','15 SEP'),('SEP','22 SEP'),('SEP','29 SEP'),
    ('OCT','6 OCT'),('OCT','13 OCT'),('OCT','20 OCT'),('OCT','27 OCT'),
    ('NOV','3 NOV'),('NOV','10 NOV'),
]
N_SEM = len(SEMANAS)

# Fases válidas para dropdown
FASES = 'PROY B,PROY E,PROY DECO,ANT.PROY,OBRA,FIN OBRA,PEDIDOS,MONTAJE,REUNIÓN,ENTREGA,APERTURA,REMATES,VIAJE,REVISIÓN,RENDERS,PDTE LIC,PENDIENTE,FUERA,CIERRE,VISITA,ALEJANDRA'
RIESGOS = 'ATENCIÓN MÁXIMA,ATENCIÓN ALTA,SEGUIMIENTO,ESTABLE,PARADO'

def make_planning_sheet(wb, nombre, dept_color, proyectos, latentes=None, supervision=None, vac=None):
    ws = wb.create_sheet(nombre)
    ws.sheet_view.showGridLines = False
    ws.sheet_view.zoomScale = 90

    # ── Cabecera departamento ─────────────────────────────────────────────────
    total_cols = 6 + N_SEM  # Nº+Proyecto+Equipo+Notas+FinEst+Riesgo + semanas
    ws.merge_cells(f'A1:{get_column_letter(total_cols)}1')
    c = ws['A1']
    c.value = f'ALEJANDRA POMBO DESIGN STUDIO  ·  {nombre.upper()}'
    c.font = fn(bold=True, color='FFFFFF', size=12)
    c.fill = fl(dept_color)
    c.alignment = al('center')
    ws.row_dimensions[1].height = 28

    # ── Fila de meses ─────────────────────────────────────────────────────────
    ws.row_dimensions[2].height = 14
    col = 7  # semanas empiezan en col 7 (A=Nº B=Proy C=Equipo D=Fase E=Notas F=FinEst G=Riesgo... wait)
    # Cols: A=Nº B=PROYECTO C=EQUIPO D=FASE_ACTUAL E=NOTAS F=FIN_EST G=RIESGO H..=semanas
    # Reasignar: A=Nº B=PROYECTO C=EQUIPO D=NOTAS E=FIN_EST F=RIESGO G..=semanas
    col_sem_start = 7  # G
    prev_mes = None
    mes_start_col = col_sem_start
    for i,(mes,sem) in enumerate(SEMANAS):
        col_i = col_sem_start + i
        if mes != prev_mes:
            if prev_mes is not None:
                end_col = col_i - 1
                if end_col > mes_start_col:
                    ws.merge_cells(start_row=2,start_column=mes_start_col,
                                   end_row=2,end_column=end_col)
                c2 = ws.cell(2, mes_start_col, mes)
                c2.font = fn(bold=True,color='FFFFFF',size=8)
                c2.fill = fl('3A3A3A')
                c2.alignment = al('center')
            mes_start_col = col_i
            prev_mes = mes
    # Last month
    end_col = col_sem_start + N_SEM - 1
    if end_col > mes_start_col:
        ws.merge_cells(start_row=2,start_column=mes_start_col,end_row=2,end_column=end_col)
    c2 = ws.cell(2, mes_start_col, prev_mes)
    c2.font = fn(bold=True,color='FFFFFF',size=8); c2.fill = fl('3A3A3A'); c2.alignment = al('center')

    # ── Cabecera columnas ─────────────────────────────────────────────────────
    ws.row_dimensions[3].height = 22
    headers = ['Nº','PROYECTO','EQUIPO','NOTAS / ESTADO ACTUAL','FIN EST.','RIESGO']
    col_widths = [6, 28, 16, 40, 10, 14]
    for i,(h,w) in enumerate(zip(headers,col_widths),1):
        c = ws.cell(3,i,h)
        c.font = fn(bold=True,color='FFFFFF',size=8)
        c.fill = fl(dept_color); c.alignment = al('center'); c.border = bd()
        ws.column_dimensions[get_column_letter(i)].width = w
    for i,(mes,sem) in enumerate(SEMANAS):
        col_i = col_sem_start + i
        c = ws.cell(3,col_i,sem)
        c.font = fn(bold=True,color='FFFFFF',size=7)
        c.fill = fl(dept_color); c.alignment = al('center'); c.border = bd()
        ws.column_dimensions[get_column_letter(col_i)].width = 10

    # ── Función para escribir una fila de proyecto ────────────────────────────
    def write_proyecto(row, codigo, nombre_p, equipo, nota, fin_est, riesgo,
                       semanas_data, alert=False, is_latente=False, is_sup=False):
        ws.row_dimensions[row].height = 20
        bg = 'FFFFFF' if row % 2 == 0 else 'F4F1EC'
        if is_latente or is_sup: bg = 'EDE8E0'

        # Nº
        c = ws.cell(row,1,codigo)
        c.font = fn(color=C['gris_l'].lstrip('FF'),size=8); c.fill = fl(bg)
        c.alignment = al('center'); c.border = bd()

        # Proyecto
        proj_color = 'C0392B' if alert else dept_color if not is_latente else '9C9088'
        c = ws.cell(row,2,nombre_p)
        c.font = fn(bold=True,color=proj_color,size=9); c.fill = fl(bg)
        c.alignment = al('left',wrap=True); c.border = bd()

        # Equipo
        c = ws.cell(row,3,equipo)
        c.font = fn(color=C['gris'].lstrip('FF'),size=8); c.fill = fl(bg)
        c.alignment = al('left'); c.border = bd()

        # Notas
        nota_corta = nota[:65] if len(nota) > 65 else nota
        c = ws.cell(row,4,nota_corta)
        c.font = fn(color=C['gris'].lstrip('FF'),size=7,italic=True); c.fill = fl(bg)
        c.alignment = al('left',wrap=True); c.border = bd()

        # Fin est
        c = ws.cell(row,5,fin_est)
        c.font = fn(color=C['gris'].lstrip('FF'),size=8); c.fill = fl(bg)
        c.alignment = al('center'); c.border = bd()

        # Riesgo
        riesgo_short = {'ATENCIÓN MÁXIMA':'⚫MÁXIMA','ATENCIÓN ALTA':'🔴ALTA',
                        'SEGUIMIENTO':'SEGUIM.','ESTABLE':'ESTABLE','PARADO':'PARADO'}.get(riesgo,riesgo)
        rcol = {'ATENCIÓN MÁXIMA':'C0392B','ATENCIÓN ALTA':'E07030',
                'SEGUIMIENTO':'1E5C1E','ESTABLE':'1E5C1E','PARADO':'9C9088'}.get(riesgo,'4A4A4A')
        c = ws.cell(row,6,riesgo_short)
        c.font = fn(bold=True,color=rcol.lstrip('FF'),size=7); c.fill = fl(bg)
        c.alignment = al('center'); c.border = bd()

        # Semanas
        for i,(mes,sem) in enumerate(SEMANAS):
            col_i = col_sem_start + i
            fase_txt = semanas_data.get(i,'') if semanas_data else ''
            fase_bg = _hex6(C.get(fase_txt, bg))
            fase_fg = 'FFFFFF' if fase_txt in FASE_TXT else '4A4A4A'
            c = ws.cell(row,col_i,fase_txt)
            c.font = fn(bold=bool(fase_txt),color=fase_fg.lstrip('FF'),size=7)
            c.fill = fl(fase_bg.lstrip('FF') if fase_bg.startswith('FF') else fase_bg)
            c.alignment = al('center'); c.border = bd()

    # ── Sección separadora ────────────────────────────────────────────────────
    def write_sep(row, label):
        ws.row_dimensions[row].height = 16
        ws.merge_cells(f'A{row}:{get_column_letter(total_cols)}{row}')
        c = ws.cell(row,1,f'  {label}')
        c.font = fn(bold=True,color=C['gris_l'].lstrip('FF'),size=8)
        c.fill = fl(C['gris_h'].lstrip('FF'))
        c.alignment = al('left')
        # Top border
        from openpyxl.styles.borders import Border, Side
        thick = Side(style='medium',color=dept_color)
        c.border = Border(top=thick)

    # ── Fila de vacaciones ────────────────────────────────────────────────────
    def write_vac(row, vac_data):
        ws.row_dimensions[row].height = 18
        c = ws.cell(row,1,'🏖')
        c.font = fn(bold=True,color='FFFFFF',size=9); c.fill = fl('2C4A2C')
        c.alignment = al('center'); c.border = bd()
        c = ws.cell(row,2,'VACACIONES')
        c.font = fn(bold=True,color='FFFFFF',size=8); c.fill = fl('2C4A2C')
        c.alignment = al('left'); c.border = bd()
        for col_i in range(3,7):
            c = ws.cell(row,col_i,'')
            c.fill = fl('2C4A2C'); c.border = bd()
        for i in range(N_SEM):
            col_i = col_sem_start + i
            txt = vac_data.get(i,'') if vac_data else ''
            c = ws.cell(row,col_i,txt)
            c.font = fn(color='FFFFFF',size=7,bold=bool(txt))
            c.fill = fl('2C4A2C'); c.alignment = al('center'); c.border = bd()

    # ── Escribir proyectos ────────────────────────────────────────────────────
    current_row = 4
    for p in proyectos:
        write_proyecto(current_row, p.get('codigo',''), p['nombre'], p.get('equipo',''),
                       p.get('nota',''), p.get('fin_est',''), p.get('riesgo','SEGUIMIENTO'),
                       p.get('semanas',{}), alert=p.get('alert',False))
        current_row += 1

    # Supervisión
    if supervision:
        write_sep(current_row, 'SUPERVISIÓN / SEGUIMIENTO PUNTUAL')
        current_row += 1
        for p in supervision:
            write_proyecto(current_row, p.get('codigo',''), p['nombre'], p.get('equipo',''),
                           p.get('nota',''), p.get('fin_est',''), p.get('riesgo','ESTABLE'),
                           p.get('semanas',{}), is_sup=True)
            current_row += 1

    # Vacaciones
    write_vac(current_row, vac or {})
    current_row += 1

    # Latentes
    if latentes:
        write_sep(current_row, 'LATENTES / PAUSADOS')
        current_row += 1
        for p in latentes:
            write_proyecto(current_row, p.get('codigo',''), p['nombre'], p.get('equipo',''),
                           p.get('nota',''), p.get('fin_est',''), p.get('riesgo','PARADO'),
                           {}, is_latente=True)
            current_row += 1

    # Freeze panes
    ws.freeze_panes = 'B4'
    return ws

# ── Semanas helper ────────────────────────────────────────────────────────────
def S(**kwargs):
    """S(s0='OBRA', s3='MONTAJE') → {0:'OBRA', 3:'MONTAJE'}"""
    return {int(k[1:]): v for k,v in kwargs.items()}

# ══════════════════════════════════════════════════════════════════════════════
# DATOS — 31 agosto 2026
# idx: 0=1SEP 1=8SEP 2=15SEP 3=22SEP 4=29SEP 5=6OCT 6=13OCT 7=20OCT 8=27OCT 9=3NOV 10=10NOV
# ══════════════════════════════════════════════════════════════════════════════

# ── RESTAURANTES ──────────────────────────────────────────────────────────────
rest = [
    dict(codigo='252',nombre='FOX + LA DESPENSA',equipo='Sofía/Naiara',
         semanas=S(s0='REMATES',s1='APERTURA'),
         nota='Amboage pdte antes 17 sep. Apertura 23 sep.',
         fin_est='23 sep',riesgo='SEGUIMIENTO'),
    dict(codigo='222',nombre='PUERTO RICO',equipo='Sofía',
         semanas=S(s5='VISITA'),fase='PROY E',
         nota='Decidir carpintero. Mediciones 2ª sem oct. Fab nov-ene. Apertura mar/abr.',
         fin_est="Mar/Abr'27",riesgo='ATENCIÓN ALTA'),
    dict(codigo='243',nombre='RUBAIYAT MADRID',equipo='Naiara',
         semanas=S(s0='ENTREGA'),
         nota='Cerrando deco terraza. Cortina, mueble, ampliación. Terminar sem sep.',
         fin_est="Sep'26",riesgo='SEGUIMIENTO'),
    dict(codigo='276',nombre='RUBAIYAT SÃO PAULO',equipo='Naiara',
         semanas=S(s3='REVISIÓN',s4='VIAJE'),fase='ANT.PROY',
         nota='Viaje 29 sep (2 noches). Anteproyecto preparando. Obra enero por fases.',
         fin_est='2027',riesgo='ATENCIÓN ALTA'),
    dict(codigo='203',nombre='BERNABÉU',equipo='Naiara',
         semanas=S(s0='CIERRE'),
         nota='Terminado. Cierre oficial: números, 1 mesa+perchas. Mesas estropeadas pdte cliente.',
         fin_est="Sep'26",riesgo='SEGUIMIENTO'),
    dict(codigo='257',nombre='MALLORCA SEVILLA',equipo='Sofía/Naiara',
         semanas=S(s0='REUNIÓN'),fase='OBRA',
         nota='Licencia concedida. Demolición hecha. Reunión sem que viene. A construir.',
         fin_est='2027',riesgo='ATENCIÓN ALTA'),
    dict(codigo='172',nombre='HIPÓDROMO',equipo='Sofía/Naiara',
         semanas={},fase='PDTE LIC',
         nota='Sin noticias. Pendiente hablar con María PM.',
         fin_est='Pdte',riesgo='SEGUIMIENTO'),
    dict(codigo='270',nombre='VIV JOANN (Anteproy.)',equipo='Sofía',
         semanas={},fase='PROY B',
         nota='Arrancando PB. Planning enviado. Obra desde 2027.',
         fin_est='2027',riesgo='SEGUIMIENTO'),
    dict(codigo='272',nombre='COBUE',equipo='Sofía/Naiara',
         semanas=S(s0='ENTREGA',s1='OBRA',s2='OBRA',s3='OBRA',s4='OBRA'),fase='OBRA',
         nota='Entrega valoración+renders esta sem. Licitando. Apertura 21 enero.',
         fin_est='21 ene',riesgo='ATENCIÓN ALTA'),
    dict(codigo='217',nombre='PASEO LAGOS 132',equipo='Sofía/Naiara',
         semanas={},fase='PROY E',
         nota='Cerrar: acabados, techos, carpintería, iluminación, domótica.',
         fin_est="May'27",riesgo='SEGUIMIENTO'),
    dict(codigo='',nombre='CARBON',equipo='Sofía',
         semanas=S(s0='REMATES'),
         nota='Remate aislamiento acústico pendiente de ejecutar.',
         fin_est="Sep'26",riesgo='SEGUIMIENTO'),
    dict(codigo='',nombre='CLÍNICA LONGEVIDAD',equipo='—',
         semanas={},
         nota='Sin novedades.',
         fin_est='—',riesgo='SEGUIMIENTO'),
]
rest_latentes = [
    dict(codigo='',nombre='SANTANDER REST',equipo='Admón.',nota='Gestión interna cobro.'),
    dict(codigo='270',nombre='JOANN DOMINICANA',equipo='Sofía',nota='Pendiente entrar.'),
    dict(codigo='',nombre='SOPHIE',equipo='Sofía',nota='Pendiente aceptación.'),
]
rest_vac = {}  # sin vacaciones en sep

# ── VIVIENDA ──────────────────────────────────────────────────────────────────
viv = [
    dict(codigo='224',nombre='PEDRAZA',equipo='Olatz/Paula',
         semanas=S(s0='REMATES'),
         nota='Reunión lun 1 sep. Lista repasos. 2ª fase caballerías pdte.',
         fin_est="Sep'26",riesgo='SEGUIMIENTO'),
    dict(codigo='',nombre='ZAHARA',equipo='Paula',
         semanas=S(s0='REMATES',s1='REMATES'),
         nota='Pedidos hechos. Dormitorio abajo pdte (post-verano).',
         fin_est="Sep'26",riesgo='SEGUIMIENTO'),
    dict(codigo='241',nombre='VALENCIA',equipo='Sara/Andrea',
         semanas=S(s1='REMATES'),
         nota='Entregado 29 ago. Mob. exterior mediados sep.',
         fin_est='29 ago',riesgo='SEGUIMIENTO'),
    dict(codigo='223',nombre='ALCALÁ 58',equipo='Alicia/Javi',
         semanas=S(s0='FIN OBRA'),fase='OBRA',
         nota='Va retrasado. Comunicar nueva fecha al cliente hoy.',
         fin_est="Sep'26",riesgo='ATENCIÓN ALTA',alert=True),
    dict(codigo='264',nombre='LA FLORIDA',equipo='Paula',
         semanas=S(s1='PROY DECO',s4='MONTAJE',s5='MONTAJE'),fase='OBRA',
         nota='Deco organizada 2ª sem sep. Montaje última sem sep.',
         fin_est="Sep'26",riesgo='ATENCIÓN ALTA'),
    dict(codigo='254',nombre='LA RINCONADA',equipo='Cristina/Nela',
         semanas=S(s8='FIN OBRA',s9='ENTREGA'),fase='OBRA',
         nota='⚠ Presu deco sin enviar — Ale URGENTE. FIN OBRA fin oct (pdte aceptación).',
         fin_est="Nov'26",riesgo='ATENCIÓN ALTA',alert=True),
    dict(codigo='271',nombre='VIUDA DE ALDAMA 3',equipo='Nela/Andrea',
         semanas=S(s0='REUNIÓN'),fase='PROY E',
         nota='Reunión jue 3 sep (Natalia+Raúl+Ale+Nela). Presupuestos pdte actualizar.',
         fin_est="Nov'26",riesgo='ATENCIÓN ALTA'),
    dict(codigo='273',nombre='VIUDA DE ALDAMA 2',equipo='Sin asignar',
         semanas={},
         nota='Nuevo proyecto. Organizar con cliente en septiembre.',
         fin_est='2027',riesgo='SEGUIMIENTO'),
    dict(codigo='216',nombre='SANTANDER VIV',equipo='Sara',
         semanas=S(s2='VISITA',s3='PROY DECO'),fase='PROY E',
         nota='Visita obra 14 sep. Reuniones deco sem 21 sep. Montaje dic.',
         fin_est="Dic'26",riesgo='ATENCIÓN ALTA'),
    dict(codigo='242',nombre='IBIZA',equipo='Sara',
         semanas=S(s1='OBRA',s2='OBRA',s3='OBRA'),fase='PDTE LIC',
         nota='Obra parada municipio. Pagos pdte clientes. Montaje diciembre.',
         fin_est="Dic'26",riesgo='SEGUIMIENTO'),
    dict(codigo='258',nombre='HERMOSILLA',equipo='Nela/Olatz',
         semanas={},fase='PENDIENTE',
         nota='Sin info. Olatz contactar al volver.',
         fin_est="Mar'27",riesgo='SEGUIMIENTO'),
    dict(codigo='268',nombre='MONTESQUINZA',equipo='Sara',
         semanas={},fase='OBRA',
         nota='⚠ Notificación licencia/DR — revisar. Obra avanzando.',
         fin_est="Mar'27",riesgo='ATENCIÓN ALTA',alert=True),
    dict(codigo='231',nombre='GAMAL STO. DOMINGO',equipo='Sofía/Olatz',
         semanas=S(s0='REMATES'),
         nota='Montaje realizado. Confirmar remates o cierre.',
         fin_est='—',riesgo='SEGUIMIENTO'),
    dict(codigo='267',nombre='CAMINO SUR 70',equipo='Cristina/Nela',
         semanas=S(s0='REMATES'),
         nota='Visita clientes 1 sep. Fallos Contraluz iluminación. Remates papel.',
         fin_est="Sep'26",riesgo='ATENCIÓN ALTA'),
    dict(codigo='211',nombre='CAMINO SUR 35',equipo='Nela',
         semanas={},fase='RENDERS',
         nota='Renders en curso (6 fases). Sin deco todavía.',
         fin_est='2027',riesgo='SEGUIMIENTO'),
    dict(codigo='239',nombre='TORRE VALENCIA',equipo='Paula',
         semanas=S(s0='REUNIÓN'),fase='PROY E',
         nota='Replanteo iluminación. Cliente sin decisiones domótica.',
         fin_est='2027',riesgo='SEGUIMIENTO'),
    dict(codigo='255',nombre='PASEO LAGOS 105',equipo='Sara/Andrea',
         semanas=S(s2='REUNIÓN'),fase='PROY B',
         nota='Reunión 15 sep. Obra retrasa a ppios 2027 (humedades).',
         fin_est='2027',riesgo='SEGUIMIENTO'),
    dict(codigo='',nombre='VALCARCE',equipo='Paula',
         semanas=S(s0='REMATES'),
         nota='Remates papel (David da largas). Remates Raúl: cliente espere 2 semanas.',
         fin_est='—',riesgo='SEGUIMIENTO'),
    dict(codigo='265',nombre='PESQUERA',equipo='Paula',
         semanas=S(s0='PEDIDOS'),fase='OBRA',
         nota='Deco en oficina pdte enviar. Constructora en marcha. Cliente: navidades.',
         fin_est="Dic'26",riesgo='SEGUIMIENTO'),
    dict(codigo='275',nombre='PASEO LAGOS 121',equipo='Olatz',
         semanas=S(s0='REUNIÓN'),fase='PROY DECO',
         nota='Reunión 2 sep. Presentación Nela — Ale revisar URGENTE.',
         fin_est="Feb'27",riesgo='ATENCIÓN ALTA',alert=True),
    dict(codigo='232',nombre='VIV PALOMA RD',equipo='Olatz/Andrea',
         semanas=S(s0='REUNIÓN'),fase='PROY DECO',
         nota='Reunión agosto con Ale. Siguiente sep.',
         fin_est="Jul'27",riesgo='SEGUIMIENTO'),
    dict(codigo='262',nombre='TEPEYAC',equipo='Paula',
         semanas={},fase='PENDIENTE',
         nota='Parado. Reunión pdte Raúl+Paula+cliente.',
         fin_est='2027',riesgo='SEGUIMIENTO'),
    dict(codigo='',nombre='ABUBILLA',equipo='Paula',
         semanas=S(s1='REMATES'),
         nota='Remates sofá y cortinas. Semana del 7 sep.',
         fin_est="Sep'26",riesgo='SEGUIMIENTO'),
]
viv_latentes = [
    dict(codigo='225',nombre='ANA HONTANAR',equipo='Paula',nota='Después de verano.'),
    dict(codigo='265',nombre='JAIME PESQUERA',equipo='Paula',nota='Después de verano.'),
    dict(codigo='',nombre='HABITACIÓN ABI',equipo='Andrea',nota='Pdte feedback presupuesto.'),
    dict(codigo='',nombre='ANTIC COLONIAL',equipo='—',nota='—'),
    dict(codigo='',nombre='MONTALBÁN',equipo='Olatz',nota='—'),
]
viv_vac = {}

# ── HOTELES ───────────────────────────────────────────────────────────────────
hot = [
    dict(codigo='70',nombre='DON RAMÓN',equipo='Marta',
         semanas=S(s0='REMATES'),
         nota='Montaje 1 sep realizado. Marta y Ale viajaron. Revisión remates.',
         fin_est='1 sep',riesgo='SEGUIMIENTO'),
    dict(codigo='185',nombre='ONE SHOT BILBAO',equipo='Marta',
         semanas=S(s6='MONTAJE',s7='MONTAJE'),fase='PROY E',
         nota='⚠ Cliente+MO cambiaron elementos. Hablar con MO. Fin oct.',
         fin_est="Oct'26",riesgo='ATENCIÓN ALTA',alert=True),
    dict(codigo='249',nombre='ROCA MAYA',equipo='Marta/Jesús',
         semanas={},fase='PROY E',
         nota='Muestras OK. Detalles listos sep→fabricar. Entrega Sem Santa: hab→ZZCC→azotea.',
         fin_est='Sem Santa',riesgo='SEGUIMIENTO'),
    dict(codigo='250',nombre='HOTEL ÚNICO BCN',equipo='Marta',
         semanas=S(s0='PROY E',s1='PROY E'),fase='PROY E',
         nota='Fichas técnicas definidas. Auna propuesto proveedor. Entregar sep y licitar.',
         fin_est="Sep'26",riesgo='SEGUIMIENTO'),
    dict(codigo='229',nombre='VÍA 66',equipo='Jesús',
         semanas=S(s0='ENTREGA',s1='PROY E',s2='PROY E'),fase='PROY E',
         nota='⚠ Entregadas 2ª+3ª plantas. MO con retrasos. ZZCC fase 2 enero.',
         fin_est="Ene'27",riesgo='ATENCIÓN ALTA',alert=True),
    dict(codigo='274',nombre='PSN PORTAL',equipo='Jesús',
         semanas=S(s0='PROY E',s1='PROY E',s2='ENTREGA',s3='OBRA',s4='OBRA'),fase='PROY E',
         nota='Entrega sep. Valoración Raúl fin sep. Obra 1 nov.',
         fin_est="Nov'26",riesgo='ATENCIÓN ALTA'),
    dict(codigo='',nombre='CASA ALMAGRO',equipo='Jesús/Marta',
         semanas=S(s0='PROY E',s1='PROY E',s2='PROY E',s3='PROY E',s4='OBRA',s5='OBRA',s6='OBRA',s7='OBRA'),fase='PROY E',
         nota='Kickoff realizado. Restaurante hotel Madrid. Obra nov/dic. ~2 meses. 200k€.',
         fin_est="Feb'27",riesgo='ATENCIÓN ALTA'),
    dict(codigo='212',nombre='HOTEL SEVILLA',equipo='Marta',
         semanas={},
         nota='Proyecto actualizado. Ampliación no aceptada. Obra casi parada.',
         fin_est='—',riesgo='SEGUIMIENTO'),
    dict(codigo='253',nombre='HOTEL GIJÓN',equipo='Marta/Jesús',
         semanas={},
         nota='Problemas instalaciones condicionan techos. Pdte últimos planos.',
         fin_est='—',riesgo='SEGUIMIENTO'),
    dict(codigo='109',nombre='VINCCI VALENCIA',equipo='Marta/Jesús',
         semanas={},
         nota='Piloto enero 2027. Montaje final jul/ago 2027.',
         fin_est="Jul'27",riesgo='SEGUIMIENTO'),
    dict(codigo='269',nombre='MARTÍNEZ CAMPOS',equipo='Jesús',
         semanas={},nota='Sin noticias.',fin_est='—',riesgo='SEGUIMIENTO'),
    dict(codigo='259',nombre='CASA KORREOS',equipo='Jesús',
         semanas={},nota='Sin noticias.',fin_est='—',riesgo='SEGUIMIENTO'),
]
hot_supervision = [
    dict(codigo='260',nombre='GONZALO CÓRDOBA',equipo='Jesús',
         semanas={},nota='Sin noticias.',fin_est='Continuo',riesgo='ESTABLE'),
    dict(codigo='235',nombre='AYALA OFICINAS',equipo='Ale/Jesús',
         semanas=S(s0='PROY DECO'),
         nota='Presu deco entregado. Montaje diciembre.',fin_est="Dic'26",riesgo='ESTABLE'),
    dict(codigo='246',nombre='KANDA SÁNCHEZ PACHECO',equipo='Jesús/Marta',
         semanas={},nota='Sin noticias.',fin_est='—',riesgo='ESTABLE'),
    dict(codigo='245',nombre='KANDA CATALINA SUÁREZ',equipo='Jesús/Marta',
         semanas={},nota='Sin noticias.',fin_est='—',riesgo='ESTABLE'),
]
hot_latentes = [
    dict(codigo='',nombre='PANTANO RURAL',equipo='—',nota='—'),
    dict(codigo='06',nombre='4 MASOS',equipo='—',nota='—'),
    dict(codigo='',nombre='BLESS',equipo='—',nota='—'),
]
hot_vac = {}

# ══════════════════════════════════════════════════════════════════════════════
# GENERAR HOJAS
# ══════════════════════════════════════════════════════════════════════════════

# Añadir fase implícita a semanas vacías
def expand_fase(p):
    fase = p.get('fase','')
    if not fase: return p
    sems = dict(p.get('semanas',{}))
    for i in range(N_SEM):
        if i not in sems:
            sems[i] = fase
    return {**p, 'semanas': sems}

rest_exp = [expand_fase(p) for p in rest]
viv_exp  = [expand_fase(p) for p in viv]
hot_exp  = [expand_fase(p) for p in hot]
hot_sup_exp = [expand_fase(p) for p in hot_supervision]

make_planning_sheet(wb, 'RESTAURANTES', '8B3A00', rest_exp, rest_latentes, vac=rest_vac)
make_planning_sheet(wb, 'VIVIENDA',     '5C1831', viv_exp,  viv_latentes,  vac=viv_vac)
make_planning_sheet(wb, 'HOTELES',      '1A3A5C', hot_exp,  hot_latentes,
                    supervision=hot_sup_exp, vac=hot_vac)

# ── Hoja de leyenda ───────────────────────────────────────────────────────────
ws_l = wb.create_sheet('LEYENDA', 0)
ws_l.sheet_view.showGridLines = False
ws_l.column_dimensions['A'].width = 20
ws_l.column_dimensions['B'].width = 30
ws_l.column_dimensions['C'].width = 15

ws_l.merge_cells('A1:C1')
c = ws_l['A1']
c.value = 'CÓDIGO DE COLORES — APDS PLANNING'
c.font = fn(bold=True,color='FFFFFF',size=12)
c.fill = fl('5C1831'); c.alignment = al('center')
ws_l.row_dimensions[1].height = 28

ws_l.cell(3,1,'FASE').font = fn(bold=True,color='FFFFFF',size=9)
ws_l.cell(3,1).fill = fl('3A3A3A'); ws_l.cell(3,1).alignment = al('center')
ws_l.cell(3,2,'DESCRIPCIÓN').font = fn(bold=True,color='FFFFFF',size=9)
ws_l.cell(3,2).fill = fl('3A3A3A')
ws_l.cell(3,3,'COLOR').font = fn(bold=True,color='FFFFFF',size=9)
ws_l.cell(3,3).fill = fl('3A3A3A'); ws_l.cell(3,3).alignment = al('center')
ws_l.row_dimensions[3].height = 18

fases_leyenda = [
    ('PROY B','Proyecto Básico en curso'),
    ('PROY E','Proyecto Ejecución en curso'),
    ('PROY DECO','Proyecto Decoración en curso'),
    ('ANT.PROY','Anteproyecto'),
    ('OBRA','En obra'),
    ('FIN OBRA','Fin de obra'),
    ('PEDIDOS','Pedidos en curso'),
    ('MONTAJE','Montaje'),
    ('ENTREGA','Entrega al cliente'),
    ('APERTURA','Apertura del negocio'),
    ('REMATES','Remates / acabados'),
    ('REUNIÓN','Reunión'),
    ('VISITA','Visita de obra'),
    ('RENDERS','Renders en curso'),
    ('VIAJE','Viaje'),
    ('ALEJANDRA','Alejandra presente'),
    ('CIERRE','Cierre de proyecto'),
    ('PDTE LIC','Pendiente licencia'),
    ('PENDIENTE','Pendiente de comenzar'),
    ('FUERA','Fuera / vacaciones'),
]
for i,(fase,desc) in enumerate(fases_leyenda):
    r = 4+i
    ws_l.row_dimensions[r].height = 16
    fase_bg = _hex6(C.get(fase,'FFEEEEEE'))
    fase_fg = 'FFFFFF' if fase in FASE_TXT else '4A4A4A'
    c = ws_l.cell(r,1,fase)
    c.font = fn(bold=True,color=fase_fg,size=9)
    c.fill = fl(fase_bg); c.alignment = al('center'); c.border = bd()
    c = ws_l.cell(r,2,desc)
    c.font = fn(size=9); c.fill = fl('F4F1EC'); c.border = bd()
    c = ws_l.cell(r,3,'')
    c.fill = fl(fase_bg); c.border = bd()

# Fecha actualización
ws_l.cell(len(fases_leyenda)+6,1,f'Actualizado: {datetime.date.today().strftime("%d/%m/%Y")}')
ws_l.cell(len(fases_leyenda)+6,1).font = fn(italic=True,color='9C9088',size=8)

out = '/home/claude/PM_ESTUDIO_MASTER_COMPLETO.xlsx'
wb.save(out)
print(f'✅ Excel guardado: {out}')
