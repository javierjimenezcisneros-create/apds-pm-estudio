#!/usr/bin/env python3
"""
APDS PM Estudio — Generador de app v3
Lee PM_ESTUDIO_MASTER_COMPLETO.xlsx y genera docs/index.html
Estructura nueva: MASTER GLOBAL + VIVIENDA/HOTELES/RESTAURANTES (Gantt sep-nov 2026)
"""
import openpyxl, json, os
from datetime import date

BASE  = os.path.dirname(os.path.abspath(__file__))
EXCEL = os.path.join(BASE, 'PM_ESTUDIO_MASTER_COMPLETO.xlsx')
OUT   = os.path.join(BASE, 'docs', 'index.html')

wb = openpyxl.load_workbook(EXCEL, data_only=True)
print("Hojas:", wb.sheetnames)

# ── Semanas: 1 sep → 10 nov 2026 (11 semanas) ────────────────────────────────
hoy = date.today()
sem_inicio = date(2026, 9, 1)
TODAY_WEEK = max(0, min((hoy - sem_inicio).days // 7, 10))
WEEK_LABELS = ['1 SEP','8 SEP','15 SEP','22 SEP','29 SEP',
               '6 OCT','13 OCT','20 OCT','27 OCT','3 NOV','10 NOV']
MONTHS = [["SEP",0,4],["OCT",5,8],["NOV",9,10]]

def find_sheet(*keywords):
    for s in wb.sheetnames:
        if any(k.upper() in s.upper() for k in keywords):
            return s
    return None

# ── MASTER GLOBAL ─────────────────────────────────────────────────────────────
# Cols: PROYECTO | DPTO | RESPONSABLE | ESTADO | HITO PRÓXIMO | FECHA |
#       BLOQUEADOR | CONSTRUCTORA | FASE | ATENCIÓN | NIVEL | OBSERVACIONES
def extract_master():
    sheet = find_sheet('MASTER','GLOBAL')
    if not sheet:
        print("⚠ No se encontró hoja MASTER GLOBAL")
        return []
    ws = wb[sheet]
    projects, header = [], False
    for row in ws.iter_rows(min_row=1, max_row=ws.max_row, max_col=12, values_only=True):
        if not any(c for c in row): continue
        first = str(row[0] or '').strip()
        # Header row detection
        if first == 'PROYECTO':
            header = True; continue
        if not header: continue
        if not row[0] or not isinstance(row[0], str): continue
        n = str(row[0]).strip()
        # Skip section separators and short strings
        if n.startswith('▸') or n.startswith('—') or n.startswith('──') or len(n) < 2: continue
        # Skip rows that look like the footer
        if 'Actualizado' in n: continue
        projects.append({
            'nombre': n,
            'dept':   str(row[1] or '').strip(),
            'resp':   str(row[2] or '—').strip(),
            'estado': str(row[3] or '—').strip(),
            'hito':   str(row[4] or '—').strip(),
            'fecha':  str(row[5] or '—').strip(),
            'bloqueador':   str(row[6] or '—').strip(),
            'constructora': str(row[7] or '—').strip(),
            'fase':    str(row[8] or '—').strip(),
            'atencion':str(row[9] or '—').strip(),
            'nivel':   str(row[10] or '—').strip(),
            'obs':     str(row[11] or '').strip() if len(row) > 11 else '',
            'tl': {}
        })
    return projects

# ── PLANNING (Gantt) ──────────────────────────────────────────────────────────
# Nueva estructura: col0=Nº, col1=PROYECTO, col2=EQUIPO, col3=NOTAS,
#                  col4=FIN EST., col5=RIESGO, col6..=semanas
def extract_planning(sheet_name):
    if not sheet_name: return {}
    ws = wb[sheet_name]
    rows = list(ws.iter_rows(min_row=1, max_row=ws.max_row,
                              max_col=ws.max_column, values_only=True))
    week_cols = {}   # col_index → week_index
    header_idx = None

    for i, row in enumerate(rows):
        # Find the header row: col1 == 'PROYECTO'
        if len(row) > 1 and str(row[1] or '').strip() == 'PROYECTO':
            header_idx = i
            # Weeks start at col 6 (0-based)
            for j in range(6, len(row)):
                c = row[j]
                if c and isinstance(c, str) and any(ch.isdigit() for ch in str(c)):
                    week_idx = j - 6   # 0=1SEP, 1=8SEP, ...
                    if week_idx < len(WEEK_LABELS):
                        week_cols[j] = week_idx
            break

    if not week_cols or header_idx is None:
        print(f"  ⚠ No se encontró cabecera en {sheet_name}")
        return {}

    tl = {}
    for row in rows[header_idx + 1:]:
        if not row or len(row) < 2: continue
        # Project name is in col 1
        n = row[1]
        if not n or not isinstance(n, str): continue
        n = n.strip()
        if len(n) < 2 or n.startswith('▸') or n.startswith('──'): continue
        if n in ('VACACIONES', '🏖', 'PROYECTO'): continue

        phases = {}
        for col, idx in week_cols.items():
            if col < len(row) and row[col] and str(row[col]).strip() not in {'None','','—'}:
                phases[idx] = str(row[col]).strip()
        if phases:
            tl[n.upper()] = phases
    return tl

# ── Extraer timelines ─────────────────────────────────────────────────────────
tl_viv  = extract_planning(find_sheet('VIVIENDA','Vivienda'))
tl_hot  = extract_planning(find_sheet('HOTELES','Hotel'))
tl_rest = extract_planning(find_sheet('RESTAURANTES','Restaurante','REST'))
all_tl  = {**tl_viv, **tl_hot, **tl_rest}

print(f"Timelines: VIV={len(tl_viv)} HOT={len(tl_hot)} REST={len(tl_rest)}")

MASTER = extract_master()

# Cruzar timelines con MASTER
for p in MASTER:
    key = p['nombre'].upper()
    for k, phases in all_tl.items():
        if k == key or k in key or key in k:
            p['tl'] = phases; break

print(f"Master: {len(MASTER)} proyectos, con timeline: {sum(1 for p in MASTER if p['tl'])}")

# ── ALERTAS derivadas del MASTER ─────────────────────────────────────────────
ALERTAS_JSON = [
    {
        'proj':  p['nombre'],
        'dept':  p['dept'],
        'resp':  p['resp'],
        'text':  p['hito'] + (' — ' + p['bloqueador']
                 if p['bloqueador'] not in ('—', '', 'None') else ''),
        'fecha': p['fecha'],
        'risk':  'CRITICO' if p['nivel'] == 'ATENCIÓN MÁXIMA' else 'ALTO'
    }
    for p in MASTER
    if p['nivel'] in ('ATENCIÓN MÁXIMA', 'ATENCIÓN ALTA')
]

# ── RESUMEN ALEJANDRA — generado desde MASTER ─────────────────────────────────
# Intentar leer de hoja si existe; si no, generar desde MASTER
ALE_ALERTAS, ALE_HITOS, ALE_MONTAJES, ALE_PRESENCIA, ALE_PROXIMAS = [], [], [], [], []
ale_sheet = find_sheet('Resumen','RESUMEN')
if ale_sheet:
    try:
        ws_ale = wb[ale_sheet]
        rows_ale = list(ws_ale.iter_rows(values_only=True))
        section = None
        for row in rows_ale:
            if not any(c for c in row): continue
            first = str(row[0] or '').strip()
            if 'ALERTAS' in first.upper(): section = 'alertas'; continue
            if 'HITOS' in first.upper(): section = 'hitos'; continue
            if 'MONTAJES' in first.upper(): section = 'montajes'; continue
            if 'PRESENCIA' in first.upper(): section = 'presencia'; continue
            if 'PRÓXIMAS' in first.upper() or 'PROXIMAS' in first.upper():
                section = 'proximas'; continue
            if first in ('NIVEL','CUÁNDO','FECHAS','FECHA','SEMANA'): continue
            if 'Alejandra Pombo' in first or first.startswith('Solo se'): continue
            if section == 'alertas' and row[0] and row[1]:
                nivel = str(row[0]).strip()
                if nivel in ('ATENCIÓN MÁXIMA','ATENCIÓN ALTA'):
                    ALE_ALERTAS.append({'nivel':nivel,'proj':str(row[1]).strip(),
                        'resp':str(row[2] or '').strip(),'text':str(row[3] or '').strip(),
                        'risk':'CRITICO' if nivel=='ATENCIÓN MÁXIMA' else 'ALTO'})
            elif section == 'hitos' and row[0] and row[1]:
                ALE_HITOS.append({'when':str(row[0]).strip(),'proj':str(row[1]).strip(),
                    'quien':str(row[2] or '').strip(),'text':str(row[3] or '').strip()})
            elif section == 'montajes' and row[0] and row[1]:
                ALE_MONTAJES.append({'fechas':str(row[0]).strip(),'proj':str(row[1]).strip(),
                    'equipo':str(row[2] or '').strip(),'notas':str(row[3] or '').strip()})
            elif section == 'presencia' and row[0] and row[1]:
                ALE_PRESENCIA.append({'fecha':str(row[0]).strip(),'texto':str(row[1]).strip()})
            elif section == 'proximas' and row[0] and row[1]:
                ALE_PROXIMAS.append({'sem':str(row[0]).strip(),'items':[str(row[1]).strip()]})
    except Exception as e:
        print(f"  ⚠ Error leyendo Resumen Alejandra: {e}")

# Si no hay hoja, derivar alertas del MASTER
if not ALE_ALERTAS:
    ALE_ALERTAS = [{'nivel':p['nivel'],'proj':p['nombre'],'resp':p['resp'],
        'text':p['hito'],'risk':'CRITICO' if p['nivel']=='ATENCIÓN MÁXIMA' else 'ALTO'}
        for p in MASTER if p['nivel'] in ('ATENCIÓN MÁXIMA','ATENCIÓN ALTA')]

print(f"Resumen Alejandra: {len(ALE_ALERTAS)} alertas, {len(ALE_HITOS)} hitos")

# ── FOCO POR PERSONA ─────────────────────────────────────────────────────────
NOMBRE_MAP = {
    'PM':'javi','Sofía':'sofia','Naiara':'naiara','Paula':'paula',
    'Sara':'sara','Olatz':'olatz','Marta':'marta','Andrea':'andrea',
    'Cristina':'cristina','Nela':'nela','Jesús':'jesus','Alejandra — CEO':'ale'
}
NOMBRE_DISPLAY = {
    'javi':'Javi','sofia':'Sofía','naiara':'Naiara','paula':'Paula','sara':'Sara',
    'olatz':'Olatz','marta':'Marta','andrea':'Andrea','cristina':'Cristina',
    'nela':'Nela','jesus':'Jesús','ale':'Alejandra'
}
PERSON_ORDER = ['javi','ale']
DEPT_SECTIONS = {'VIVIENDA':[],'RESTAURANTES':[],'HOTELES':[]}
PROJS_PERSONA, DEPT_PERSONA, NOMBRE_PERSONA, FOCO_SEMANA = {}, {}, {}, {}

foco_sheet = find_sheet('FOCO','PERSONA')
if foco_sheet:
    try:
        ws_foco = wb[foco_sheet]
        current_dept = None
        for row in ws_foco.iter_rows(min_row=2, values_only=True):
            if not row[0]: continue
            nombre_excel = str(row[0]).strip()
            if nombre_excel.startswith('──'):
                if 'VIVIENDA' in nombre_excel.upper(): current_dept = 'VIVIENDA'
                elif 'RESTAURANTE' in nombre_excel.upper(): current_dept = 'RESTAURANTES'
                elif 'HOTEL' in nombre_excel.upper(): current_dept = 'HOTELES'
                continue
            pid = NOMBRE_MAP.get(nombre_excel)
            if not pid: continue
            dept_raw = str(row[1] or '').strip()
            if 'VIVIENDA' in dept_raw: dept = 'Vivienda'
            elif 'HOTEL' in dept_raw: dept = 'Hoteles'
            elif 'RESTAURANTE' in dept_raw: dept = 'Restaurantes'
            else: dept = 'todos'
            projs = [p.strip().upper() for p in str(row[2] or '').replace('·','|').split('|')
                     if p.strip() and len(p.strip()) > 2]
            foco_items = [f.strip() for f in str(row[3] or '').split('·')
                          if f.strip() and len(f.strip()) > 4]
            NOMBRE_PERSONA[pid] = NOMBRE_DISPLAY.get(pid, nombre_excel)
            DEPT_PERSONA[pid] = dept
            PROJS_PERSONA[pid] = projs[:8]
            FOCO_SEMANA[pid] = foco_items[:6]
            if pid not in ('javi','ale'):
                target = current_dept or ('VIVIENDA' if dept=='Vivienda' else
                          'RESTAURANTES' if dept=='Restaurantes' else 'HOTELES')
                if target in DEPT_SECTIONS:
                    DEPT_SECTIONS[target].append(pid)
        for dept_pids in DEPT_SECTIONS.values():
            PERSON_ORDER.extend(dept_pids)
    except Exception as e:
        print(f"  ⚠ Error leyendo Foco por persona: {e}")

# Si no hay hoja de foco, derivar personas del MASTER
if not NOMBRE_PERSONA:
    # Generar foco básico desde responsables en MASTER
    seen_pid = set()
    for p in MASTER:
        for resp_raw in str(p.get('resp','')).replace('/',' ').replace('+',' ').split():
            pid = NOMBRE_MAP.get(resp_raw.strip())
            if pid and pid not in seen_pid and pid not in ('javi','ale'):
                seen_pid.add(pid)
                NOMBRE_PERSONA[pid] = NOMBRE_DISPLAY.get(pid, resp_raw)
                DEPT_PERSONA[pid] = p.get('dept','').capitalize()
                PROJS_PERSONA[pid] = []
                FOCO_SEMANA[pid] = []
                dept = p.get('dept','').upper()
                if dept in DEPT_SECTIONS:
                    DEPT_SECTIONS[dept].append(pid)
    for dept_pids in DEPT_SECTIONS.values():
        PERSON_ORDER.extend(dept_pids)

print(f"Personas: {PERSON_ORDER}")

# ── RENDERS ───────────────────────────────────────────────────────────────────
RENDERS = []
renders_sheet = find_sheet('Render','render','RENDER')
if renders_sheet:
    try:
        ws_r = wb[renders_sheet]
        header = False
        for row in ws_r.iter_rows(values_only=True):
            if not any(c for c in row): continue
            first = str(row[0] or '').strip()
            if first == 'PRIO': header = True; continue
            if not header: continue
            if not row[0] or str(row[0]).strip().startswith(('▸','──')): continue
            prio_raw = str(row[0]).strip()
            prio = 'URGENTE' if '🔴' in prio_raw else 'MEDIO' if '🟠' in prio_raw else 'BAJO'
            proj = str(row[1] or '').strip()
            if not proj or len(proj) < 2: continue
            RENDERS.append({
                'prio':prio,'proj':proj,'dept':str(row[2] or '').strip(),
                'resp':str(row[3] or '').strip(),'renderista':str(row[4] or 'Challan').strip(),
                'fase':str(row[5] or '').strip(),'estado':str(row[6] or '').strip(),
                'comprometida':str(row[8] or '').strip() if len(row)>8 else '',
                'realista':str(row[9] or '').strip() if len(row)>9 else '',
                'next':str(row[10] or '').strip() if len(row)>10 else '',
                'riesgo':'CRITICO' if prio=='URGENTE' else 'MEDIO'
            })
    except Exception as e:
        print(f"  ⚠ Error leyendo Renders: {e}")

RENDERS_FLUJO = [
    {'fase':'F1','titulo':'Volumetría y capturas','desc':'Planos + 3D + vistas marcadas. ~1 semana desde doc completa.'},
    {'fase':'F2','titulo':'Materialidad y acabados','desc':'Toda la info de acabados y mobiliario. 2-3 días.'},
    {'fase':'F3','titulo':'Render final','desc':'TODAS las correcciones cerradas ANTES de renderizar.'},
]
RENDERS_REGLAS = [
    'Máx. 2-3 proyectos activos simultáneos por renderista.',
    'Nunca activar sin documentación 100% completa.',
    'PM es el filtro — nadie pasa trabajo sin visto bueno del PM.',
]

MONTAJES = ALE_MONTAJES if ALE_MONTAJES else []
AGENDA = [{'when':a['when'],'text':f"{a['proj']} — {a['text']}"} for a in ALE_HITOS[:8]]
VAC = {}

# ── GENERAR DATA JS ───────────────────────────────────────────────────────────
DATA_JS = f"""
const WEEK_LABELS = {json.dumps(WEEK_LABELS)};
const MONTHS = {json.dumps(MONTHS)};
const TODAY_WEEK = {TODAY_WEEK};
const WEEK_LABEL = 'Semana {hoy.strftime("%d %b %Y")}';
const WEEK_STAMP = 'SEM {hoy.strftime("%d-%m-%Y")}';
const UPDATED = 'Actualizado {hoy.strftime("%d %b %Y")}';
const MASTER = {json.dumps(MASTER, ensure_ascii=False)};
const ALERTAS = {json.dumps(ALERTAS_JSON, ensure_ascii=False)};
const AGENDA = {json.dumps(AGENDA, ensure_ascii=False)};
const MONTAJES = {json.dumps(MONTAJES, ensure_ascii=False)};
const ALE_ALERTAS = {json.dumps(ALE_ALERTAS, ensure_ascii=False)};
const ALE_HITOS = {json.dumps(ALE_HITOS, ensure_ascii=False)};
const ALE_MONTAJES = {json.dumps(ALE_MONTAJES, ensure_ascii=False)};
const ALE_PRESENCIA = {json.dumps(ALE_PRESENCIA, ensure_ascii=False)};
const ALE_PROXIMAS = {json.dumps(ALE_PROXIMAS, ensure_ascii=False)};
const RENDERS = {json.dumps(RENDERS, ensure_ascii=False)};
const RENDERS_FLUJO = {json.dumps(RENDERS_FLUJO, ensure_ascii=False)};
const RENDERS_REGLAS = {json.dumps(RENDERS_REGLAS, ensure_ascii=False)};
const RENDERS_VAC = '';
const FOCO_SEMANA = {json.dumps(FOCO_SEMANA, ensure_ascii=False)};
const PROJS_PERSONA = {json.dumps(PROJS_PERSONA, ensure_ascii=False)};
const DEPT_PERSONA = {json.dumps(DEPT_PERSONA, ensure_ascii=False)};
const NOMBRE_PERSONA = {json.dumps(NOMBRE_PERSONA, ensure_ascii=False)};
const VAC = {json.dumps(VAC, ensure_ascii=False)};
const DEPT_SECTIONS = {json.dumps(DEPT_SECTIONS, ensure_ascii=False)};
const PERSON_ORDER = {json.dumps(PERSON_ORDER)};
"""

# ── CONSTRUIR HTML ────────────────────────────────────────────────────────────
import re

def escape_script_close(s):
    return s.replace('</script', '<\\/script')

head_html = open(os.path.join(BASE, 'head.html'), encoding='utf-8').read()
app_js    = open(os.path.join(BASE, 'app.js'),   encoding='utf-8').read()

html_out = (head_html +
    '<script>\n' + escape_script_close(DATA_JS) + '\n</script>\n' +
    '<script>\n' + app_js + '\n</script>\n' +
    '</body>\n</html>')

os.makedirs(os.path.dirname(OUT), exist_ok=True)
with open(OUT, 'w', encoding='utf-8') as f:
    f.write(html_out)

print(f"\n✅ App generada: {OUT}")
print(f"   Proyectos: {len(MASTER)} | Alertas: {len(ALERTAS_JSON)} | Con timeline: {sum(1 for p in MASTER if p['tl'])}")
